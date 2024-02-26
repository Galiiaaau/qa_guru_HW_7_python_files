from zipfile import ZipFile
from pypdf import PdfReader
import csv
from io import TextIOWrapper
from openpyxl import load_workbook

from tests.conftest import resource_zip_path


def test_pdf():
    with ZipFile(resource_zip_path) as zip_file:
        with zip_file.open('PDF.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            assert 'засмеялся' in reader.pages[100].extract_text()


def test_csv():
    with ZipFile(resource_zip_path) as zip_file:
        with zip_file.open('CSV.csv') as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'windows-1251'), delimiter=';'))
            fourth_row = csvreader[3]
            print(fourth_row[3])
            assert "OU002" == fourth_row[1]
            assert "Екатеринбург" in fourth_row[3]


def test_xlsx():
    with ZipFile(resource_zip_path) as zip_file:
        with zip_file.open('XLSX.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            name = sheet.cell(row=2, column=6).value
            print(name)
            assert name == "Руководитель департамента"
